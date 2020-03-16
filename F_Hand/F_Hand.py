#Read from txt and write as it is to other txt
#error - charmap' codec can't decode byte 0x8d in position 3252: character maps to <undefined> Solution-use encoding="UTF-8"
#readline()-Returns one line from the file
#readlines()-Returns a list of lines from the file
chatfile=open("E:\Intellipaat\Python\DevOps_CloudBabies.txt","rt",encoding="UTF-8")
for line in chatfile:
    print(line)
    writefile=open("E:\Intellipaat\Python\DevOps_CloudBabiesWrite.txt","a",encoding="UTF-8")
    writefile.write(line)
chatfile.close()
writefile.close()

#Read from txt and write to other txt with 2 fields
import re
chatfile=open("E:\Intellipaat\Python\DevOps_CloudBabies.txt","rt",encoding="UTF-8")
writefile=open("E:\Intellipaat\Python\DevOps_CloudBabiesDtMb.txt","w",encoding="UTF-8")
for line in chatfile:
    writefile=open("E:\Intellipaat\Python\DevOps_CloudBabiesDtMb.txt","a",encoding="UTF-8")
    #validate mm/dd/yy format - here yy=10 to 99
    dateFormat=re.findall("(0[1-9]|[12][0-9]|3[01])[- /.](0[1-9]|1[012])[- /.]([1-9][0-9])",line)
    mbNumDetails=re.findall("[+]\d+\s*.+\s*\d*[-\s]\d*:",line)
    if(dateFormat and mbNumDetails):
        date=str(dateFormat[0][0])+"/"+str(dateFormat[0][1])+"/"+str(dateFormat[0][2])
        for mbNumDetail in mbNumDetails:
            if(mbNumDetail):
                mbNum=re.split(":",mbNumDetail,1)
        writefile.write(date+"     "+mbNum[0]+"\n")
        #print(date,mbNum[0])
    elif(dateFormat):
        date=str(dateFormat[0][0])+"/"+str(dateFormat[0][1])+"/"+str(dateFormat[0][2])
        writefile.write(date+"\n")
        #print(date)
chatfile.close()
writefile.close()

#Read from txt and write to other CSV
#csv library built into Python, and see how CSV parsing works using the pandas library
#write to a CSV file using a writer object and the .write_row() method
import re
import csv
chatfile=open("E:\Intellipaat\Python\DevOps_CloudBabies.txt","rt",encoding="UTF-8")
#newline='' else will add new line after evert write row
with open('E:\Intellipaat\Python\DevOps_CloudBabiesDtMb.csv',mode='w',newline='') as dtMbCSVfile:
    #delimiter=',', quotechar='"',quoting=csv.QUOTE_MINIMAL these are optional field
    csvfile_writer=csv.writer(dtMbCSVfile,delimiter=',',quotechar='"',quoting=csv.QUOTE_MINIMAL)
    csvfile_writer.writerow(['Date','Mobile Num'])
    for line in chatfile:
    #validate mm/dd/yy format - here yy=10 to 99
        dateFormat=re.findall("(0[1-9]|[12][0-9]|3[01])[- /.](0[1-9]|1[012])[- /.]([1-9][0-9])",line)
        mbNumDetails=re.findall("[+]\d+\s*.+\s*\d*[-\s]\d*:",line)
        if(dateFormat and mbNumDetails):
            date=str(dateFormat[0][0])+"/"+str(dateFormat[0][1])+"/"+str(dateFormat[0][2])
            for mbNumDetail in mbNumDetails:
                if(mbNumDetail):
                    mbNum=re.split(":",mbNumDetail,1)
            csvfile_writer.writerow([date,mbNum[0]])
            #print(date,mbNum[0])
        elif(dateFormat):
            date=str(dateFormat[0][0])+"/"+str(dateFormat[0][1])+"/"+str(dateFormat[0][2])
            csvfile_writer.writerow([date])
            #print(date)
chatfile.close()
dtMbCSVfile.close()

#Read from CSV and write to CSV
import csv
with open('E:\Intellipaat\Python\DevOps_CloudBabiesDtMb.csv') as csvReadfile:
    csv_reader=csv.reader(csvReadfile, delimiter=',')
    with open('E:\Intellipaat\Python\DevOps_CloudBabiesDtMbcopy.csv',mode='w',newline='') as csvWriteFile:
        csv_writer=csv.writer(csvWriteFile,delimiter=',')
        for row in csv_reader:
            csv_writer.writerow(row)   
            #print(row)
csvReadfile.close()
csvWriteFile.close()

#Read from excel and write to other Excel
#Openpyxl is a Python library for reading and writing Excel (with extension xlsx/xlsm/xltx/xltm) files.
#The openpyxl module allows Python program to read and modify Excel files.
import openpyxl
from openpyxl import Workbook
#path for to read a file
path="E:\Intellipaat\Python\DevOps_CloudBabiesDtMb_x.xlsx"
#load workbook then take active sheet
wb_obj=openpyxl.load_workbook(path)
sheet_obj=wb_obj.active
maxrow=sheet_obj.max_row
maxcol=sheet_obj.max_column
#create new workbook
wb=Workbook()
sheet=wb.active
#print(maxrow)
#print(maxcol)
for rowval in range(1,maxrow+1):
    #print("rowval",rowval)
    for colval in range (1,maxcol+1):
        #read from excel
        cell_obj=sheet_obj.cell(row=rowval,column=colval)
        #write to excel
        cell_wr=sheet.cell(row=rowval,column=colval)
        cell_wr.value=cell_obj.value
wb.save("E:\Intellipaat\Python\DevOps_CloudBabiesDtMb_x_Copy.xlsx")      

# Reading from txt and write to excel
import re
import openpyxl
chatfile=open("E:\Intellipaat\Python\DevOps_CloudBabies.txt","rt",encoding="UTF-8")
wb=Workbook()
sheet=wb.active
rownum=1
colnum=1
for line in chatfile:
    #validate mm/dd/yy format - here yy=10 to 99
    dateFormat=re.findall("(0[1-9]|[12][0-9]|3[01])[- /.](0[1-9]|1[012])[- /.]([1-9][0-9])",line)
    mbNumDetails=re.findall("[+]\d+\s*.+\s*\d*[-\s]\d*:",line)
    if(dateFormat and mbNumDetails):
        date=str(dateFormat[0][0])+"/"+str(dateFormat[0][1])+"/"+str(dateFormat[0][2])
        for mbNumDetail in mbNumDetails:
            if(mbNumDetail):
                mbNum=re.split(":",mbNumDetail,1)
        #write to excel
        cell_wr_date=sheet.cell(row=rownum,column=colnum)
        cell_wr_mbNum=sheet.cell(row=rownum,column=colnum+1)
        cell_wr_date.value=date
        cell_wr_mbNum.value=mbNum[0]
    elif(dateFormat):
        date=str(dateFormat[0][0])+"/"+str(dateFormat[0][1])+"/"+str(dateFormat[0][2])
        #write to excel
        cell_wr_date=sheet.cell(row=rownum,column=colnum)
        cell_wr_date.value=date
    rownum=rownum+1
wb.save(("E:\Intellipaat\Python\DevOps_CloudBabiesDtMb_txt_XLS.xlsx") )
chatfile.close()